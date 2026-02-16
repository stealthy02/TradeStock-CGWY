from typing import Dict, Any, List
from datetime import datetime
import os
import sys
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from io import BytesIO
from pathlib import Path

# 检测是否在 Electron 打包环境中运行
def get_base_dir():
    # 如果在打包环境中，使用 sys._MEIPASS 获取打包后的资源目录
    if hasattr(sys, '_MEIPASS'):
        return Path(sys._MEIPASS)
    # 否则使用文件所在目录的上级目录
    return Path(__file__).parent.parent.parent




def _format_value(value):
    """
    格式化值，处理日期等特殊类型
    """
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def convert_to_xlsx(data: Dict[str, Any], bill_type: str = "purchase") -> bytes:
    """
    将对账单数据转换为xlsx文件流
    
    Args:
        data: 对账单数据，包含bill_info、purchase_list/sale_list、pay_record_list/receipt_list
        bill_type: 对账单类型，"purchase"或"sale"
    
    Returns:
        xlsx文件的字节流
    """
    # 获取基础目录
    base_dir = get_base_dir()
    
    # 确定模板文件路径
    template_base_dir = base_dir / "public"
    if bill_type == "purchase":
        template_dir = template_base_dir / "purchase"
        name_key = "supplier_name"
        list_key = "purchase_list"
    else:
        template_dir = template_base_dir / "sale"
        name_key = "purchaser_name"
        list_key = "sale_list"
    
    # 获取名称
    entity_name = data["bill_info"].get(name_key, "")
    
    # 查找模板文件
    template_files = []
    if template_dir.exists():
        template_files = [f.name for f in template_dir.iterdir() if f.suffix == ".xlsx"]
    
    # 选择模板文件
    template_path = None
    if entity_name:
        # 查找特定模板
        specific_template = f"{entity_name}_模板.xlsx"
        if specific_template in template_files:
            template_path = template_dir / specific_template
    
    # 如果没有特定模板，使用通用模板
    if not template_path:
        generic_template = "通用模板.xlsx"
        if generic_template in template_files:
            template_path = template_dir / generic_template
    
    # 如果没有模板文件，返回空字节流
    if not template_path:
        return b""
    
    # 加载模板
    wb = load_workbook(template_path)
    ws = wb.active
    
    # 第一步：替换基本信息字段
    bill_info = data["bill_info"]
    
    # 处理日期字段：如果end_date为空，使用今日日期
    start_date = bill_info.get("start_date")
    end_date = bill_info.get("end_date")
    if not end_date or end_date == "":
        end_date = datetime.now().strftime("%Y-%m-%d")
    
    replace_mapping = {
        "{{supplier_name}}": _format_value(bill_info.get("supplier_name")),
        "{{purchaser_name}}": _format_value(bill_info.get("purchaser_name")),
        "{{start_date}}": _format_value(start_date),
        "{{end_date}}": _format_value(end_date),
        "{{bill_amount}}": _format_value(bill_info.get("bill_amount")),
        "{{statement_amount}}": _format_value(bill_info.get("statement_amount"))
    }
    
    # 遍历所有单元格，替换文本
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell_value = str(cell.value)
                for placeholder, value in replace_mapping.items():
                    if placeholder in cell_value:
                        cell.value = cell_value.replace(placeholder, value)
            elif ws.merged_cells.ranges and any(cell.coordinate in r for r in ws.merged_cells.ranges):
                # 处理合并单元格：获取合并区域的左上角单元格的值
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                        if top_left_cell.value:
                            cell_value = str(top_left_cell.value)
                            for placeholder, value in replace_mapping.items():
                                if placeholder in cell_value:
                                    top_left_cell.value = cell_value.replace(placeholder, value)
                        break
    
    # 第二步：处理表格数据
    list_data = data.get(list_key, {}).get("list", [])
    if list_data:
        # 查找<list_data>所在行
        list_data_row = None
        fixed_rows = None
        for idx, row in enumerate(ws.iter_rows(), 1):
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value)
                    match = re.search(r'<list_data(:\d+)?>', cell_value)
                    if match:
                        list_data_row = idx
                        if match.group(1):
                            fixed_rows = int(match.group(1)[1:])
                        # 删除<list_data>标记
                        cell.value = cell_value.replace(match.group(0), "")
                        break
            if list_data_row:
                break
        
        # 如果找到<list_data>行，处理表格数据
        if list_data_row:
            # 确定需要的行数
            if fixed_rows:
                needed_rows = max(fixed_rows, len(list_data))
            else:
                needed_rows = len(list_data)
            
            # 保存原始模板行的所有单元格信息（包括样式）
            template_cells = {}
            for cell in ws[list_data_row]:
                template_cells[cell.column] = {
                    "value": cell.value,
                    "font": cell.font,
                    "border": cell.border,
                    "fill": cell.fill,
                    "alignment": cell.alignment,
                    "number_format": cell.number_format
                }
            
            # 先插入需要的行数（从第2行开始插入）
            for i in range(needed_rows - 1):
                ws.insert_rows(list_data_row + i + 1)
            
                # 填充数据到每一行
            for i in range(needed_rows):
                current_row = list_data_row + i
                
                # 获取当前行的数据
                if i < len(list_data):
                    row_data = list_data[i]
                else:
                    row_data = {}
                
                # 填充单元格数据
                for col, template_info in template_cells.items():
                    target_cell = ws.cell(row=current_row, column=col)
                    
                    # 复制样式
                    if template_info["font"]:
                        target_cell.font = template_info["font"].copy()
                    if template_info["border"]:
                        target_cell.border = template_info["border"].copy()
                    if template_info["fill"]:
                        target_cell.fill = template_info["fill"].copy()
                    if template_info["alignment"]:
                        target_cell.alignment = template_info["alignment"].copy()
                    if template_info["number_format"]:
                        target_cell.number_format = template_info["number_format"]
                    
                    # 处理单元格值
                    if template_info["value"]:
                        cell_value = str(template_info["value"])
                        
                        # 替换字段
                        field_mapping = {
                            "{{product_name}}": _format_value(row_data.get("product_name")),
                            "{{customer_product_name}}": _format_value(row_data.get("customer_product_name")),
                            "{{purchase_date}}": _format_value(row_data.get("purchase_date")),
                            "{{sale_date}}": _format_value(row_data.get("sale_date")),
                            "{{total_num}}": _format_value(row_data.get("total_num")),
                            "{{total_kg}}": _format_value(row_data.get("total_kg")),
                            "{{total_price}}": _format_value(row_data.get("total_price")),
                            "{{product_spec}}": _format_value(row_data.get("product_spec")),
                            "{{remark}}": _format_value(row_data.get("remark")),
                            "{{unit_price}}": _format_value(row_data.get("unit_price"))
                        }
                        
                        # 替换占位符
                        for placeholder, value in field_mapping.items():
                            if placeholder in cell_value:
                                cell_value = cell_value.replace(placeholder, value)
                        
                        target_cell.value = cell_value
                    else:
                        target_cell.value = None
    
    # 保存为字节流
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()