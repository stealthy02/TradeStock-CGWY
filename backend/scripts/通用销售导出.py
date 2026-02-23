from typing import Dict, Any
from datetime import datetime
import os
import sys
import re
import csv
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from io import BytesIO
from pathlib import Path


def get_base_dir():
    """获取正确的基础目录"""
    if hasattr(sys, '_MEIPASS'):
        return Path(sys._MEIPASS)
    # 开发环境：从脚本所在目录向上两级
    return Path(__file__).parent.parent


def _format_value(value):
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def load_product_units() -> Dict[str, str]:
    """
    加载商品单位映射
    从脚本同级目录的商品单位.csv文件读取
    
    Returns:
        商品名称到数量单位的映射字典
    """
    units_map = {}
    
    try:
        # 获取脚本所在目录
        script_dir = Path(__file__).parent
        csv_path = script_dir / "商品单位.csv"
        
        if csv_path.exists():
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    product_name = row.get("name", "").strip()
                    unit = row.get("unit", "").strip()
                    if product_name and unit:
                        units_map[product_name] = unit
    except Exception as e:
        pass
    
    return units_map


def export(data: Dict[str, Any]) -> bytes:
    base_dir = get_base_dir()
    
    # 加载商品单位映射
    product_units = load_product_units()
    
    # 优先检查 exe 所在目录是否有 public 文件夹
    exe_dir = None
    if hasattr(sys, 'executable') and sys.executable:
        exe_dir = Path(sys.executable).parent
        external_public = exe_dir / "public"
        if external_public.exists():
            base_dir = exe_dir
    
    template_base_dir = base_dir / "public" / "sale"
    
    entity_name = data["bill_info"].get("purchaser_name", "")
    
    template_files = []
    if template_base_dir.exists():
        template_files = [f.name for f in template_base_dir.iterdir() if f.suffix == ".xlsx"]
    
    template_path = None
    if entity_name:
        specific_template = f"{entity_name}_模板.xlsx"
        if specific_template in template_files:
            template_path = template_base_dir / specific_template
    
    if not template_path:
        generic_template = "通用模板.xlsx"
        if generic_template in template_files:
            template_path = template_base_dir / generic_template
    
    if not template_path:
        return b""
    
    wb = load_workbook(template_path)
    ws = wb.active
    
    bill_info = data["bill_info"]
    
    start_date = bill_info.get("start_date")
    end_date = bill_info.get("end_date")
    if not end_date or end_date == "":
        end_date = datetime.now().strftime("%Y-%m-%d")
    
    replace_mapping = {
        "{{supplier_name}}": _format_value(bill_info.get("supplier_name")),
        "{{purchaser_name}}": _format_value(bill_info.get("purchaser_name")),
        "{{start_date}}": _format_value(start_date),
        "{{end_date}}": _format_value(end_date),
        "{{bill_amount}}": _format_value(bill_info.get("statement_amount")),
        "{{statement_amount}}": _format_value(bill_info.get("statement_amount"))
    }
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell_value = str(cell.value)
                for placeholder, value in replace_mapping.items():
                    if placeholder in cell_value:
                        cell.value = cell_value.replace(placeholder, value)
            elif ws.merged_cells.ranges and any(cell.coordinate in r for r in ws.merged_cells.ranges):
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                        if top_left_cell.value:
                            cell_value = str(top_left_cell.value)
                            for placeholder, value in replace_mapping.items():
                                if placeholder in cell_value:
                                    top_left_cell.value = cell_value.replace(placeholder, value)
                        break
    
    list_data = data.get("sale_list", {}).get("list", [])
    if list_data:
        list_data_row = None
        fixed_rows = None
        for idx, row in enumerate(ws.iter_rows(), 1):
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value)
                    match = re.search(r'<list_data(:\d+)?>', cell_value)
                    if match:
                        list_data_row = idx
                        if match.group(1):
                            fixed_rows = int(match.group(1)[1:])
                        cell.value = cell_value.replace(match.group(0), "")
                        break
            if list_data_row:
                break
        
        if list_data_row:
            if fixed_rows:
                needed_rows = max(fixed_rows, len(list_data))
            else:
                needed_rows = len(list_data)
            
            template_cells = {}
            for cell in ws[list_data_row]:
                template_cells[cell.column] = {
                    "value": cell.value,
                    "font": cell.font,
                    "border": cell.border,
                    "fill": cell.fill,
                    "alignment": cell.alignment,
                    "number_format": cell.number_format
                }
            
            for i in range(needed_rows - 1):
                ws.insert_rows(list_data_row + i + 1)
            
            for i in range(needed_rows):
                current_row = list_data_row + i
                
                if i < len(list_data):
                    row_data = list_data[i]
                else:
                    row_data = {}
                
                # 获取商品数量单位
                product_name = row_data.get("product_name", "").strip()
                unit = product_units.get(product_name, "")
                
                for col, template_info in template_cells.items():
                    target_cell = ws.cell(row=current_row, column=col)
                    
                    if template_info["font"]:
                        target_cell.font = template_info["font"].copy()
                    if template_info["border"]:
                        target_cell.border = template_info["border"].copy()
                    if template_info["fill"]:
                        target_cell.fill = template_info["fill"].copy()
                    if template_info["alignment"]:
                        target_cell.alignment = template_info["alignment"].copy()
                    if template_info["number_format"]:
                        target_cell.number_format = template_info["number_format"]
                    
                    if template_info["value"]:
                        cell_value = str(template_info["value"])
                        
                        field_mapping = {
                            "{{product_name}}": _format_value(row_data.get("product_name")),
                            "{{customer_product_name}}": _format_value(row_data.get("customer_product_name")),
                            "{{sale_date}}": _format_value(row_data.get("sale_date")),
                            "{{total_num}}": _format_value(row_data.get("total_num")),
                            "{{total_kg}}": _format_value(row_data.get("total_kg")),
                            "{{total_price}}": _format_value(row_data.get("total_price")),
                            "{{product_spec}}": _format_value(row_data.get("product_spec")),
                            "{{remark}}": _format_value(row_data.get("remark")),
                            "{{unit_price}}": _format_value(row_data.get("unit_price")),
                            "{{unit}}": _format_value(unit)
                        }
                        
                        for placeholder, value in field_mapping.items():
                            if placeholder in cell_value:
                                cell_value = cell_value.replace(placeholder, value)
                        
                        target_cell.value = cell_value
                    else:
                        target_cell.value = None
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
